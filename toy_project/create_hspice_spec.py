#!/usr/bin/env python3
"""Generate HSPICE MC data extraction specification Excel for real PDK."""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Color / style definitions ──
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill("solid", fgColor="D6E4F0")
SUBHEADER_FONT = Font(bold=True, size=10)
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
REQUIRED_FILL = PatternFill("solid", fgColor="E2EFDA")
OPTIONAL_FILL = PatternFill("solid", fgColor="FCE4EC")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

def style_header_row(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

def style_data_area(ws, start_row, end_row, max_col):
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")

def auto_width(ws, max_col, min_width=12, max_width=50):
    for c in range(1, max_col + 1):
        best = min_width
        for row in ws.iter_rows(min_col=c, max_col=c, values_only=False):
            for cell in row:
                if cell.value:
                    lines = str(cell.value).split("\n")
                    longest = max(len(l) for l in lines)
                    best = max(best, min(longest + 2, max_width))
        ws.column_dimensions[get_column_letter(c)].width = best

# ======================================================================
# Sheet 1: 개요 및 지침 (Overview & Instructions)
# ======================================================================
ws0 = wb.active
ws0.title = "개요 및 지침"
ws0.merge_cells("A1:F1")
ws0.cell(row=1, column=1, value="HSPICE MC Simulation Data Extraction Specification — SRAM Vmin Surrogate Modeling").font = Font(bold=True, size=14)

info_rows = [
    ("프로젝트 목적", "Real PDK SRAM bitcell의 HSPICE Monte Carlo simulation 결과를 추출하여\nGP surrogate model + differentiable physics layer 기반 Vmin 추정 및 inverse design pipeline 구축"),
    ("데이터 사용처", "추출된 data는 Python (PyTorch) 기반 GP surrogate model의 training/validation에 사용됨"),
    ("예상 simulation 규모", "~300개 (common_N, PU) 조건 × 6개 Vop level × 1,000~5,000 MC samples = 총 ~1,800만 ~ 9,000만 MC simulation"),
    ("예상 disk 용량", "각 MC condition 당 ~1-5MB (CSV 기준) → 총 ~300MB ~ 1.5GB (선택적 압축 가능)"),
    ("Z_target 의미", "Vmin = Zscore(SNMR) = 6.0을 만족하는 Vdd. 64Mb SRAM @ 99.9% 수율 기준.\n실제 PDK target에 따라 조정 가능 (Sheet2 참조)"),
    ("데이터 형식 우선순위", "1순위: CSV 파일 (Python pandas로 바로 loading)\n2순위: .npz (압축, 용량 작음)\n3순위: raw HSPICE .tr0/.sw0 (parsing 필요, 비권장)"),
    ("주의사항", "• 모든 MC simulation은 동일한 seed (seed=42)로 시작하여 reproducibility 확보\n• simulation 실패 시 해당 condition을 반드시 logging (NaN으로 표시)\n• 사내 PDK 보안: PDK 이름/파라미터는 코드 내에서만 참조, 외부 유출 금지"),
]

ws0.cell(row=3, column=1, value="항목").font = Font(bold=True)
ws0.cell(row=3, column=2, value="내용").font = Font(bold=True)
for i, (k, v) in enumerate(info_rows, start=4):
    ws0.cell(row=i, column=1, value=k)
    ws0.cell(row=i, column=2, value=v)
    ws0.cell(row=i, column=1).font = Font(bold=True)

auto_width(ws0, 2, min_width=20, max_width=80)
ws0.column_dimensions["B"].width = 100

# ======================================================================
# Sheet 2: 고정 파라미터 (Fixed Parameters)
# ======================================================================
ws1 = wb.create_sheet("고정 파라미터")
headers1 = ["파라미터", "값/범위", "단위", "필수여부", "설명/근거"]
for c, h in enumerate(headers1, 1):
    ws1.cell(row=1, column=c, value=h)
style_header_row(ws1, 1, len(headers1))

fixed_params = [
    ("Technology node", "<PDK명 입력>", "nm", "필수", "사용 중인 PDK (예: SKY130, TSMC 28nm, Samsung 14nm 등)"),
    ("SRAM cell type", "6T SRAM bitcell", "-", "필수", "표준 6T SRAM (PU/PD/PG)"),
    ("SNM type", "Hold SNM (SNMR)", "-", "필수", "Hold static noise margin (right-hand side). Read SNM은 별도로 필요시 추출."),
    ("Temperature", "25 (상온)", "°C", "필수", "단일 온도. 추후 확장 시 -40~125°C"),
    ("VDD (supply)", "NOM_VDD ± 10% (예: 1.8V)", "V", "필수", "공정 nominal voltage 기준. SUR MARGIN 분석을 위해 ±10% range 필요."),
    ("Vop levels", "[0.4, 0.5, 0.6, 0.7, 0.8, 0.9]", "V", "필수", "Vmin interpolation을 위한 6개 level. toy project와 동일.\n만약 PDK의 min Vdd가 0.6V 이상이면 [0.6, 0.7, 0.8, 0.9, 1.0, 1.1]로 조정."),
    ("Z_target (Vmin 기준)", "6.0", "sigma", "권장", "Vmin 정의를 위한 target Z-score.\nZ = mu_SNMR / sigma_SNMR = 6.0 → 64Mb @ 99.9% 수율.\n실제 product spec에 따라 조정 가능."),
    ("MC sample size", "1,000 ~ 5,000", "samples", "필수", "각 (common_N, PU, Vop) condition 당 MC sample 수.\n적을수록 sigma 추정의 분산 증가.\n권장: 1,000 (minimum), 3,000 (권장), 5,000 (high accuracy)"),
    ("MC seed", "42", "-", "필수", "reproducibility 보장. 모든 condition에서 동일 seed 사용."),
    ("Process corner (baseline)", "TT (Typical-Typical)", "-", "필수", "기본 분석은 TT corner. 추후 FF/SS/FS/SF extension 가능."),
    ("Global variation model", "<PDK 변수명>", "-", "필수", "NMOS/PMOS global Vth variation parameter name (PDK-specific).\n예: `mismatch=2` for local, `global_variation=on` for global."),
    ("common_N shift 구현", "<PDK 변수명>", "mV", "필수", "NMOS (PG+PD)의 Vth를 일괄 shift하는 parameter.\nPDK에 따라 `tnom` / `delvto` / `vth_shift_n` 등.\n+10mV = slower NMOS."),
    ("PU shift 구현", "<PDK 변수명>", "mV", "필수", "PMOS (PU)의 Vth를 shift하는 parameter.\nPDK에 따라 `delvto_p` / `vth_shift_p` 등.\n+10mV = slower PMOS (higher |Vth|)."),
]

for i, row_data in enumerate(fixed_params, start=2):
    for c, val in enumerate(row_data, 1):
        cell = ws1.cell(row=i, column=c, value=val)
        if c == 4 and val == "필수":
            cell.fill = REQUIRED_FILL
        elif c == 4 and val == "권장":
            cell.fill = OPTIONAL_FILL

style_data_area(ws1, 2, len(fixed_params) + 1, len(headers1))
auto_width(ws1, len(headers1), min_width=15, max_width=60)

# ======================================================================
# Sheet 3: Simulation Matrix (Sweep Parameters)
# ======================================================================
ws2 = wb.create_sheet("Simulation Matrix")
headers2 = ["common_N_shift (mV)", "PU_shift (mV)", "Vop (V)", "MC samples", "Priority", "비고"]
for c, h in enumerate(headers2, 1):
    ws2.cell(row=1, column=c, value=h)
style_header_row(ws2, 1, len(headers2))

# Build full factorial design
cn_values = [-60, -40, -20, -10, 0, 10, 20, 40, 60]  # 9 levels
pu_values = [-60, -40, -20, -10, 0, 10, 20, 40, 60]  # 9 levels
vop_values = [0.4, 0.5, 0.6, 0.7, 0.8, 0.9]  # 6 levels

row_idx = 2
idx = 0
for cn in cn_values:
    for pu in pu_values:
        # Determine priority
        if cn in (-10, 0, 10) and pu in (-10, 0, 10):
            priority = "P1 (core region)"
        elif abs(cn) <= 20 and abs(pu) <= 20:
            priority = "P2 (near-nominal)"
        elif cn in (-60, 60) or pu in (-60, 60):
            priority = "P3 (corner/extreme)"
        else:
            priority = "P2 (near-nominal)"

        for vop in vop_values:
            note = ""
            if cn == 0 and pu == 0 and vop == 0.6:
                note = "★ baseline condition (reference)"
            ws2.cell(row=row_idx, column=1, value=cn)
            ws2.cell(row=row_idx, column=2, value=pu)
            ws2.cell(row=row_idx, column=3, value=vop)
            ws2.cell(row=row_idx, column=4, value="3,000")
            ws2.cell(row=row_idx, column=5, value=priority)
            ws2.cell(row=row_idx, column=6, value=note)
            row_idx += 1
            idx += 1

# Summary
row_idx += 2
ws2.cell(row=row_idx, column=1, value="요약").font = Font(bold=True, size=12)
row_idx += 1
n_cn = len(cn_values)
n_pu = len(pu_values)
n_vop = len(vop_values)
n_total = n_cn * n_pu * n_vop
summary_data = [
    ("common_N levels", str(n_cn), f"{cn_values[0]} ~ {cn_values[-1]} mV"),
    ("PU levels", str(n_pu), f"{pu_values[0]} ~ {pu_values[-1]} mV"),
    ("Vop levels", str(n_vop), "0.4 ~ 0.9 V"),
    ("총 simulation 조건", f"{n_total:,}", f"= {n_cn} × {n_pu} × {n_vop}"),
    ("MC/condition", "3,000", "권장값. 시간/자원에 따라 1,000~5,000 조정"),
    ("총 MC simulation 횟수", f"{n_total * 3000:,}", f"= {n_total:,} × 3,000"),
]
for i, (k, v, note) in enumerate(summary_data):
    ws2.cell(row=row_idx, column=1, value=k).font = Font(bold=True)
    ws2.cell(row=row_idx, column=2, value=v)
    ws2.cell(row=row_idx, column=3, value=note)
    row_idx += 1

style_data_area(ws2, 2, row_idx, len(headers2))
auto_width(ws2, len(headers2), min_width=15, max_width=40)

# ======================================================================
# Sheet 4: 출력 데이터 형식 (Output Data Format)
# ======================================================================
ws3 = wb.create_sheet("출력 데이터 형식")

# Section A: Per-condition format
ws3.cell(row=1, column=1, value="A. 각 simulation condition 당 출력 형식 (CSV)").font = Font(bold=True, size=12)
ws3.merge_cells("A1:H1")

headers3a = ["Column", "이름", "타입", "예시", "단위", "필수", "설명"]
for c, h in enumerate(headers3a, 1):
    ws3.cell(row=2, column=c, value=h)
style_header_row(ws3, 2, len(headers3a))

col_defs = [
    ("1", "common_N_shift", "float", "0.0", "mV", "필수", "NMOS (PG+PD) Vth shift amount"),
    ("2", "PU_shift", "float", "0.0", "mV", "필수", "PMOS (PU) Vth shift amount"),
    ("3", "vop", "float", "0.6", "V", "필수", "Supply voltage level for this MC run"),
    ("4", "mc_sample_id", "int", "0", "-", "필수", "MC sample index (0 ~ N-1)"),
    ("5", "snmr", "float", "0.245", "V", "필수", "Hold SNM (right-hand) value from HSPICE"),
    ("6", "snml", "float", "0.238", "V", "선택", "Hold SNM (left-hand) value — 있으면 유용"),
    ("7", "iread", "float", "1.2e-6", "A", "선택", "Read current — 있으면 추가 분석 가능"),
    ("8", "temperature", "float", "25.0", "°C", "권장", "Simulation temperature"),
    ("9", "process_corner", "str", "TT", "-", "권장", "Process corner label"),
    ("10", "status", "str", "ok/fail", "-", "필수", "Simulation convergence status"),
]

for i, row_data in enumerate(col_defs, start=3):
    for c, val in enumerate(row_data, 1):
        cell = ws3.cell(row=i, column=c, value=val)
        if c == 6:
            if val == "필수":
                cell.fill = REQUIRED_FILL
            elif val == "권장":
                cell.fill = NOTE_FILL

style_data_area(ws3, 3, 2 + len(col_defs), len(headers3a))

# Section B: Aggregated format (preferred)
row_start = 3 + len(col_defs) + 2
ws3.cell(row=row_start, column=1, value="B. 집계(Aggregated) 출력 형식 — 권장").font = Font(bold=True, size=12)
ws3.merge_cells(f"A{row_start}:H{row_start}")

headers3b = ["Column", "이름", "타입", "예시", "단위", "필수", "설명"]
r = row_start + 1
for c, h in enumerate(headers3b, 1):
    ws3.cell(row=r, column=c, value=h)
style_header_row(ws3, r, len(headers3b))

agg_defs = [
    ("1", "common_N_shift", "float", "0.0", "mV", "필수", "NMOS Vth shift"),
    ("2", "PU_shift", "float", "0.0", "mV", "필수", "PMOS Vth shift"),
    ("3", "vop", "float", "0.6", "V", "필수", "Supply voltage"),
    ("4", "n_samples", "int", "3000", "-", "필수", "Number of valid MC samples"),
    ("5", "mu_snmr", "float", "0.237", "V", "필수", "Mean of SNMR across MC samples"),
    ("6", "sigma_snmr", "float", "0.012", "V", "필수", "Std dev of SNMR across MC samples"),
    ("7", "min_snmr", "float", "0.185", "V", "권장", "Minimum SNMR (worst-case)"),
    ("8", "max_snmr", "float", "0.289", "V", "권장", "Maximum SNMR"),
    ("9", "skewness_snmr", "float", "-0.15", "-", "선택", "Skewness of SNMR distribution"),
    ("10", "kurtosis_snmr", "float", "3.2", "-", "선택", "Kurtosis (excess) of SNMR dist."),
    ("11", "n_fail", "int", "0", "-", "권장", "Number of failed MC samples"),
    ("12", "temperature", "float", "25.0", "°C", "권장", "Temperature"),
    ("13", "corner", "str", "TT", "-", "권장", "Process corner"),
    ("14", "file_path", "str", "results/cond_000.csv", "-", "선택", "원본 raw data file path"),
]

r2 = r + 1
for i, row_data in enumerate(agg_defs, start=r2):
    for c, val in enumerate(row_data, 1):
        cell = ws3.cell(row=i, column=c, value=val)
        if c == 6:
            if val == "필수":
                cell.fill = REQUIRED_FILL
            elif val == "권장":
                cell.fill = NOTE_FILL

style_data_area(ws3, r2, r2 + len(agg_defs) - 1, len(headers3b))
auto_width(ws3, len(headers3b), min_width=14, max_width=50)

# ======================================================================
# Sheet 5: Hold-out Validation Data
# ======================================================================
ws4 = wb.create_sheet("Hold-out Validation")

ws4.cell(row=1, column=1, value="Hold-out (검증) 데이터 사양").font = Font(bold=True, size=12)
ws4.merge_cells("A1:F1")

val_headers = ["항목", "값/설명", "필수여부"]
for c, h in enumerate(val_headers, 1):
    ws4.cell(row=2, column=c, value=h)
style_header_row(ws4, 2, len(val_headers))

val_data = [
    ("목적", "GP surrogate model의 Vmin 예측 정확도를 평가하기 위한 ground-truth 데이터.\nSurrogate training에는 사용하지 않고 오직 validation에만 사용.", "필수"),
    ("Data size", "전체 데이터의 ~20% (약 10~20개 (common_N, PU) condition)", "필수"),
    ("선정 방식", "(common_N, PU) 2D 공간에서 균일하게 분포하도록 stratified sampling.\n극단 corner (FS, SF)와 near-nominal region 포함.", "필수"),
    ("Vmin ground-truth 측정법", "각 (common_N, PU) holdout condition에 대해:\n  1. 모든 Vop level에서 MC simulation → mu, sigma\n  2. compute_vmin_from_z()와 동일한 방식으로 Vmin 계산\n  3. toy project의 physics layer와 동일한 알고리즘 사용", "필수"),
    ("Hold-out condition list", "아래 'Hold-out 조건 목록' 참조.\n총 16개 조건 (전체 81개의 ~20%)", "필수"),
    ("출력 형식", "Sheet4 (B)와 동일한 aggregated format.\n단, 'is_holdout' column = True로 표시.", "필수"),
]

for i, (k, v, req) in enumerate(val_data, start=3):
    ws4.cell(row=i, column=1, value=k).font = Font(bold=True)
    ws4.cell(row=i, column=2, value=v)
    ws4.cell(row=i, column=3, value=req)
    if req == "필수":
        ws4.cell(row=i, column=3).fill = REQUIRED_FILL

# Hold-out condition list
row_ho = 3 + len(val_data) + 1
ws4.cell(row=row_ho, column=1, value="Hold-out 조건 목록 (추천)").font = Font(bold=True, size=11)
ws4.merge_cells(f"A{row_ho}:F{row_ho}")

ho_headers = ["common_N_shift (mV)", "PU_shift (mV)", "Region", "선정 이유"]
for c, h in enumerate(ho_headers, 1):
    ws4.cell(row=row_ho + 1, column=c, value=h)
style_header_row(ws4, row_ho + 1, len(ho_headers))

# 16 holdout conditions (evenly distributed, covering all quadrants)
holdout_conds = [
    (-40, -40, "FF-like", "Corner: both fast"),
    (-40, 0, "N-fast nominal-P", "Edge"),
    (-40, 40, "FS-like (fast N, slow P)", "Corner"),
    (-20, -20, "Near FF", "Mid edge"),
    (-20, 20, "Mid", "Mid"),
    (-20, 40, "Edge", "Edge"),
    (0, -40, "Nominal N, fast P", "Edge"),
    (0, 0, "Nominal", "★ Center — 가장 중요"),
    (0, 40, "Nominal N, slow P", "Edge"),
    (20, -40, "Edge", "Edge"),
    (20, -20, "Mid", "Mid"),
    (20, 40, "Near SS", "Mid edge"),
    (40, -40, "SF-like (slow N, fast P)", "Corner"),
    (40, 0, "N-slow nominal-P", "Edge"),
    (40, 40, "SS-like", "Corner: both slow"),
    (-60, 60, "Extreme FS", "극단 corner test"),
]

for i, (cn, pu, region, reason) in enumerate(holdout_conds, start=row_ho + 2):
    ws4.cell(row=i, column=1, value=cn)
    ws4.cell(row=i, column=2, value=pu)
    ws4.cell(row=i, column=3, value=region)
    ws4.cell(row=i, column=4, value=reason)

style_data_area(ws4, row_ho + 2, row_ho + 1 + len(holdout_conds), len(ho_headers))
auto_width(ws4, 4, min_width=15, max_width=60)

# ======================================================================
# Sheet 6: 확장 파라미터 (Extensions — Optional)
# ======================================================================
ws5 = wb.create_sheet("확장 (선택사항)")

ws5.cell(row=1, column=1, value="확장 시뮬레이션 (선택사항 — 시간/자원이 허락될 때만)").font = Font(bold=True, size=12)
ws5.merge_cells("A1:E1")

ext_headers = ["확장 항목", "설명", "추가 조건 수", "필요 MC samples", "우선순위"]
for c, h in enumerate(ext_headers, 1):
    ws5.cell(row=2, column=c, value=h)
style_header_row(ws5, 2, len(ext_headers))

extensions = [
    ("추가 temperature", "-40°C, 125°C 추가\n(현재 25°C만)", "81 × 2 = 162", "~500,000", "중간"),
    ("FF/SS corner", "Process corner FF, SS에서 동일한\nsimulation matrix 반복", "81 × 6 × 2 = 972", "~3,000,000", "낮음 (paper 위해 필요)"),
    ("FS/SF corner", "FS, SF corner 추가", "81 × 6 × 2 = 972", "~3,000,000", "낮음"),
    ("Read SNM", "SNMW (read SNM) 추가 측정.\nVmin estimation의 second metric.", "Same matrix", "Same", "선택"),
    ("Mb (memory bank) sweep", "Mb size를 변수로 추가.\nVmin = f(common_N, PU, Mb)", "81 × 6 × 3 = 1,458", "~4,400,000", "낮음 (차기 연구)"),
    ("Write margin", "Write margin 측정 추가.\nWrite Vmin 별도 추정.", "Same matrix", "Same", "선택"),
    ("Local mismatch 분리", "Global variation 외 local mismatch\n영향을 별도로 추출", "Same matrix", "5,000+/cond", "낮음"),
]

for i, (name, desc, n_cond, n_mc, pri) in enumerate(extensions, start=3):
    ws5.cell(row=i, column=1, value=name).font = Font(bold=True)
    ws5.cell(row=i, column=2, value=desc)
    ws5.cell(row=i, column=3, value=n_cond)
    ws5.cell(row=i, column=4, value=n_mc)
    ws5.cell(row=i, column=5, value=pri)

style_data_area(ws5, 3, 2 + len(extensions), len(ext_headers))
auto_width(ws5, len(ext_headers), min_width=15, max_width=50)

# ======================================================================
# Sheet 7: HSPICE Netlist Template
# ======================================================================
ws6 = wb.create_sheet("HSPICE Netlist Template")

ws6.cell(row=1, column=1, value="HSPICE Netlist 구조 (참고용 — PDK에 맞게 수정 필요)").font = Font(bold=True, size=12)
ws6.merge_cells("A1:C1")

tmpl_headers = ["Section", "내용", "비고"]
for c, h in enumerate(tmpl_headers, 1):
    ws6.cell(row=2, column=c, value=h)
style_header_row(ws6, 2, len(tmpl_headers))

template = [
    (".OPTION", "POST=2 PROBE RUNLVL=6", "Run level 6 = maximum accuracy for MC"),
    (".TEMP", "25", "Temperature (고정)"),
    (".PARAM", "common_N_shift=0.0  PU_shift=0.0", "Sweep parameter — values from Sheet 3"),
    (".PARAM", "vop=0.6", "Supply voltage — sweep across Vop levels"),
    ("* PDK include", ".lib '<PDK_path>/xxx.lib' TT", "TT corner library. Replace with PDK path."),
    ("* SRAM cell netlist", "<6T SRAM bitcell subcircuit>", "PDK-specific SRAM cell"),
    (".PARAM", "vth_shift_n='common_N_shift*1e-3'", "★ NMOS Vth shift implementation.\nPDK parameter name에 맞게 수정 필수"),
    (".PARAM", "vth_shift_p='PU_shift*1e-3'", "★ PMOS Vth shift implementation.\nPDK parameter name에 맞게 수정 필수"),
    ("* MC block", ".DATA MC_DATA\n+  ...\n.ENDDATA", "MC sweep definition"),
    (".MC", "MC_SNMR MONTE=3000\n+ ...", "3,000 MC samples. 결과를 .tr0에 저장."),
    (".PROBE", "TRAN v(q) v(qb)", "Internal node voltages for SNM measurement"),
    ("* SNM measurement", ".MEASURE SNMR ...\n+ DERIVED='...'", "Hold SNM measurement (PDK-specific method)"),
    (".ALTER", "case_1 / case_2 / ...", "각 (common_N, PU) condition을 ALTER로 구현,\n또는 별도 netlist 파일로 분리"),
    ("출력 파일", "cond_{idx}.mc0 / cond_{idx}.tr0", "조건별 출력. 자동화를 위해\nPython script로 netlist 생성 권장"),
]

for i, (sec, desc, note) in enumerate(template, start=3):
    ws6.cell(row=i, column=1, value=sec).font = Font(bold=True)
    ws6.cell(row=i, column=2, value=desc)
    ws6.cell(row=i, column=3, value=note)

style_data_area(ws6, 3, 2 + len(template), len(tmpl_headers))
auto_width(ws6, len(tmpl_headers), min_width=30, max_width=80)

# ======================================================================
# Sheet 8: 체크리스트 (Checklist)
# ======================================================================
ws7 = wb.create_sheet("체크리스트")

ws7.cell(row=1, column=1, value="Data 추출 전/후 체크리스트").font = Font(bold=True, size=12)
ws7.merge_cells("A1:D1")

chk_headers = ["#", "항목", "확인", "비고"]
for c, h in enumerate(chk_headers, 1):
    ws7.cell(row=2, column=c, value=h)
style_header_row(ws7, 2, len(chk_headers))

checks = [
    ("1", "PDK library path 정확한지 확인", "☐", "Sheet2 참조"),
    ("2", "NMOS/PMOS Vth shift parameter name 확인", "☐", "PDK 문서 필수 확인"),
    ("3", "MC seed=42로 고정", "☐", "reproducibility"),
    ("4", "TT corner에서만 우선 추출 (확장은 선택)", "☐", "기본은 TT only"),
    ("5", "총 81 (common_N,PU) × 6 (Vop) = 486개 조건", "☐", "Sheet3 matrix 확인"),
    ("6", "각 조건 MC 3,000 samples (최소 1,000)", "☐", "자원에 따라 조정"),
    ("7", "Simulation 실패 condition logging", "☐", "Sheet4 `status` column"),
    ("8", "Hold-out 16개 condition 별도 표시", "☐", "Sheet5 목록 참조"),
    ("9", "출력 형식: CSV (권장) 또는 .npz", "☐", "Sheet4 (B) format"),
    ("10", "파일 naming: cond_{idx:03d}.csv", "☐", "idx = 0~485"),
    ("11", "Data integrity check: mu/sigma range sanity", "☐", "Python loading 후 검증"),
    ("12", "사내 PDK data 보안 조치 확인", "☐", "PDK명/경로 코드 내에서만 사용"),
]

for i, (num, item, chk, note) in enumerate(checks, start=3):
    ws7.cell(row=i, column=1, value=num)
    ws7.cell(row=i, column=2, value=item)
    ws7.cell(row=i, column=3, value=chk)
    ws7.cell(row=i, column=4, value=note)

style_data_area(ws7, 3, 2 + len(checks), len(chk_headers))
auto_width(ws7, len(chk_headers), min_width=10, max_width=60)
ws7.column_dimensions["B"].width = 55

# ── Save ──
out_path = Path(__file__).resolve().parent / "HSPICE_Data_Extraction_Spec.xlsx"
wb.save(str(out_path))
print(f"Spec saved: {out_path}")
